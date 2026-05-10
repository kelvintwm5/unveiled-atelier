"""
Generates Bridal Business Budget 2025 XLSX for Google Sheets import.
"""

import base64
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Colour palette ──────────────────────────────────────────────────────────
GREY_BG   = "F3F3F3"
BLUE_BG   = "D9E8FB"   # alternating row tint for Transactions
WHITE_BG  = "FFFFFF"
HEADER_FONT = Font(bold=True)

def header_fill():
    return PatternFill("solid", fgColor=GREY_BG)

def blue_fill():
    return PatternFill("solid", fgColor=BLUE_BG)

def bold(ws, cell_ref):
    ws[cell_ref].font = Font(bold=True)

def apply_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font  = HEADER_FONT
        cell.fill  = header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")

def freeze(ws, cell="A2"):
    ws.freeze_panes = ws[cell]

SGD_FORMAT = '"S$"#,##0.00'
PCT_FORMAT  = '0.00%'


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – Transactions
# ══════════════════════════════════════════════════════════════════════════════
TRANSACTIONS = [
    # Date       Type      Description                      Category                  Amount   Payment        Client             Receipt   Notes
    ("2025-01-08","Income","Bridal gown package – Sarah Lim","Gown Sales",           3800.00,"PayNow",       "Sarah Lim",       "REC-001","Full payment"),
    ("2025-01-12","Expense","Fabric & lace materials",       "Cost of Goods",          520.00,"Bank Transfer","",                "EXP-001","For Feb stock"),
    ("2025-01-18","Income","Gown alteration – Mrs Chan",    "Alteration Services",    280.00,"Cash",         "Mrs Chan",        "REC-002","Hem & bustle"),
    ("2025-01-25","Expense","Instagram ads – Jan",           "Marketing & Advertising",150.00,"Bank Transfer","",                "EXP-002","Monthly boost"),
    ("2025-02-03","Income","Rental deposit – Amanda Koh",   "Deposits Received",      500.00,"PayNow",       "Amanda Koh",      "REC-003","Deposit for Apr wedding"),
    ("2025-02-10","Income","Veil & accessories – Priya S",  "Accessories",            320.00,"Cash",         "Priya Subramaniam","REC-004","Pearl veil set"),
    ("2025-02-14","Expense","Studio rental – Feb",           "Rental & Utilities",     800.00,"Bank Transfer","",                "EXP-003","Monthly studio"),
    ("2025-02-22","Expense","Alterations – external tailor", "Alterations & Tailoring",180.00,"Cash",         "",                "EXP-004","Subcontract"),
    ("2025-03-05","Income","Gown rental – Grace Tan",       "Gown Rental",            650.00,"PayNow",       "Grace Tan",       "REC-005","3-day rental"),
    ("2025-03-12","Expense","Garment bags & packaging",      "Packaging & Supplies",    95.00,"Cash",         "",                "EXP-005","Bulk order"),
    ("2025-03-20","Income","Bridal gown – Michelle Ng",     "Gown Sales",            4200.00,"Bank Transfer","Michelle Ng",     "REC-006","Includes fitting"),
    ("2025-03-28","Expense","Sewing machine service",        "Equipment & Tools",      120.00,"Cash",         "",                "EXP-006","Annual service"),
    ("2025-04-05","Income","Alteration bundle – 3 gowns",   "Alteration Services",    480.00,"PayNow",       "Various",         "REC-007","Walk-in clients"),
    ("2025-04-15","Expense","Accounting fees – Q1",          "Professional Services",  350.00,"Bank Transfer","",                "EXP-007","Quarterly bookkeeping"),
    ("2025-05-02","Income","Gown sales – Rebecca Ho",       "Gown Sales",            3500.00,"PayNow",       "Rebecca Ho",      "REC-008","Customised A-line"),
]

def build_transactions(wb):
    ws = wb.active
    ws.title = "Transactions"
    ws.sheet_properties.tabColor = "4472C4"   # blue

    headers = [
        "Date","Type","Description","Category",
        "Amount (SGD)","Payment Method","Client Name","Receipt No.","Notes"
    ]
    ws.append(headers)
    apply_header_row(ws, 1)
    freeze(ws)

    for i, row in enumerate(TRANSACTIONS, start=2):
        ws.append(list(row))
        # alternating tint
        fill = blue_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE_BG)
        for cell in ws[i]:
            cell.fill = fill
        # currency format on Amount column (E)
        ws.cell(row=i, column=5).number_format = SGD_FORMAT

    # column widths (approximate)
    col_widths = [13, 9, 38, 26, 14, 16, 22, 12, 25]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – Monthly Summary
# ══════════════════════════════════════════════════════════════════════════════
MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

def build_monthly_summary(wb):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = "70AD47"   # green

    headers = ["Month","Total Income (SGD)","Total Expenses (SGD)",
               "Net Profit (SGD)","Profit Margin (%)"]
    ws.append(headers)
    apply_header_row(ws, 1)
    freeze(ws)

    for i, month in enumerate(MONTHS, start=2):
        month_abbr = month[:3].upper()   # e.g. "JAN"
        # SUMIFS: match month abbreviation against Transactions date column
        # We compare TEXT(Transactions!A:A,"MMM") to the first 3 chars of month
        month_num = i - 1   # 1-12

        income_formula  = (
            f'=SUMPRODUCT((MONTH(Transactions!$A$2:$A$1000)={month_num})'
            f'*(YEAR(Transactions!$A$2:$A$1000)=2025)'
            f'*(Transactions!$B$2:$B$1000="Income")'
            f'*Transactions!$E$2:$E$1000)'
        )
        expense_formula = (
            f'=SUMPRODUCT((MONTH(Transactions!$A$2:$A$1000)={month_num})'
            f'*(YEAR(Transactions!$A$2:$A$1000)=2025)'
            f'*(Transactions!$B$2:$B$1000="Expense")'
            f'*Transactions!$E$2:$E$1000)'
        )
        net_formula     = f'=B{i}-C{i}'
        margin_formula  = f'=IF(B{i}=0,0,D{i}/B{i})'

        ws.append([month, income_formula, expense_formula, net_formula, margin_formula])
        ws[f'B{i}'].number_format = SGD_FORMAT
        ws[f'C{i}'].number_format = SGD_FORMAT
        ws[f'D{i}'].number_format = SGD_FORMAT
        ws[f'E{i}'].number_format = PCT_FORMAT

    # Totals row
    total_row = 14
    ws.append(["TOTAL",
               f'=SUM(B2:B13)',
               f'=SUM(C2:C13)',
               f'=SUM(D2:D13)',
               f'=IF(B{total_row}=0,0,D{total_row}/B{total_row})'])
    for cell in ws[total_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill()
    ws[f'B{total_row}'].number_format = SGD_FORMAT
    ws[f'C{total_row}'].number_format = SGD_FORMAT
    ws[f'D{total_row}'].number_format = SGD_FORMAT
    ws[f'E{total_row}'].number_format = PCT_FORMAT

    col_widths = [14, 22, 24, 20, 18]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – IRAS 4-Line Statement
# ══════════════════════════════════════════════════════════════════════════════
EXPENSE_CATEGORIES = [
    "Cost of Goods",
    "Alterations & Tailoring",
    "Rental & Utilities",
    "Marketing & Advertising",
    "Professional Services",
    "Transport",
    "Packaging & Supplies",
    "Equipment & Tools",
    "Training & Development",
    "Bank & Payment Fees",
    "Miscellaneous",
]

def build_iras(wb):
    ws = wb.create_sheet("IRAS 4-Line Statement")
    ws.sheet_properties.tabColor = "ED7D31"   # orange

    # Year selector
    ws["A1"] = "Year"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = 2025
    ws["B1"].font = Font(bold=True)

    ws.append([])  # blank row 2

    # ── 4-Line Statement ────────────────────────────────────────────────────
    ws["A3"] = "IRAS 4-Line Statement"
    ws["A3"].font = Font(bold=True, size=13)

    headers_4line = ["Line","Description","Amount (SGD)"]
    ws.append(headers_4line)   # row 4
    apply_header_row(ws, 4)

    # Revenue = all Income rows for selected year
    revenue_formula = (
        '=SUMPRODUCT((YEAR(Transactions!$A$2:$A$1000)=$B$1)'
        '*(Transactions!$B$2:$B$1000="Income")'
        '*Transactions!$E$2:$E$1000)'
    )
    # Expenses = all Expense rows for selected year
    expense_formula = (
        '=SUMPRODUCT((YEAR(Transactions!$A$2:$A$1000)=$B$1)'
        '*(Transactions!$B$2:$B$1000="Expense")'
        '*Transactions!$E$2:$E$1000)'
    )

    four_lines = [
        (1, "Revenue (Total Receipts)",              revenue_formula),
        (2, "Gross Profit/Loss (service business – same as Revenue)", '=C5'),
        (3, "Allowable Business Expenses",           f'={expense_formula[1:]}'),  # same formula
        (4, "Adjusted Profit/Loss (Net Profit)",     '=C5-C7'),
    ]

    for r_offset, (line, desc, formula) in enumerate(four_lines, start=5):
        ws.append([line, desc, formula])
        ws.cell(row=r_offset, column=3).number_format = SGD_FORMAT
        if line == 4:
            ws.cell(row=r_offset, column=1).font = Font(bold=True)
            ws.cell(row=r_offset, column=2).font = Font(bold=True)
            ws.cell(row=r_offset, column=3).font = Font(bold=True)

    ws.append([])  # blank row 9

    # ── Expense breakdown ──────────────────────────────────────────────────
    ws["A10"] = "Expense Breakdown by Category"
    ws["A10"].font = Font(bold=True, size=12)

    ws.append(["Category", "", "Amount (SGD)"])  # row 11
    apply_header_row(ws, 11)

    for i, cat in enumerate(EXPENSE_CATEGORIES, start=12):
        cat_formula = (
            f'=SUMPRODUCT((YEAR(Transactions!$A$2:$A$1000)=$B$1)'
            f'*(Transactions!$B$2:$B$1000="Expense")'
            f'*(Transactions!$D$2:$D$1000="{cat}")'
            f'*Transactions!$E$2:$E$1000)'
        )
        ws.append([cat, "", cat_formula])
        ws.cell(row=i, column=3).number_format = SGD_FORMAT

    # Total expenses check row
    total_row = 12 + len(EXPENSE_CATEGORIES)
    ws.append(["Total Expenses", "", f'=SUM(C12:C{total_row-1})'])
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=3).number_format = SGD_FORMAT
    ws.cell(row=total_row, column=3).font = Font(bold=True)

    freeze(ws, "A2")
    col_widths = [52, 5, 18]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – Records Log
# ══════════════════════════════════════════════════════════════════════════════
def build_records_log(wb):
    ws = wb.create_sheet("Records Log")
    ws.sheet_properties.tabColor = "A5A5A5"   # grey

    ws["A1"] = "NOTE: IRAS requires all business records to be kept for 5 years from the Year of Assessment."
    ws["A1"].font = Font(bold=True, italic=True, color="C00000")
    ws.merge_cells("A1:E1")

    ws.append([])  # row 2 blank

    headers = ["Month","Total Transactions","Receipts Saved (Y/N)",
               "Storage Location","Notes"]
    ws.append(headers)   # row 3
    apply_header_row(ws, 3)
    ws.freeze_panes = "A4"

    for i, month in enumerate(MONTHS, start=4):
        month_num = i - 3   # 1-12
        tx_count_formula = (
            f'=COUNTPRODUCT((MONTH(Transactions!$A$2:$A$1000)={month_num})'
            f'*(YEAR(Transactions!$A$2:$A$1000)=2025)'
            f'*(Transactions!$A$2:$A$1000<>""))'
        )
        # Simpler COUNTIFS alternative (works in Sheets)
        tx_count_formula = (
            f'=COUNTIFS(Transactions!$A$2:$A$1000,"<>",'
            f'Transactions!$A$2:$A$1000,">="&DATE(2025,{month_num},1),'
            f'Transactions!$A$2:$A$1000,"<="&EOMONTH(DATE(2025,{month_num},1),0))'
        )
        ws.append([month, tx_count_formula, "", "Google Drive / Physical Folder", ""])

    col_widths = [13, 20, 22, 30, 25]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    build_transactions(wb)
    build_monthly_summary(wb)
    build_iras(wb)
    build_records_log(wb)

    out_path = "/tmp/bridal_budget_2025.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")

    with open(out_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    print(f"Base64 length: {len(b64)}")
    print("DONE")
    return b64

if __name__ == "__main__":
    main()
